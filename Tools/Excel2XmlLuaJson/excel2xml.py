from genericpath import exists
import openpyxl
import os

def excel2xml(excel_path, xml_path):
    book = openpyxl.load_workbook(excel_path)
    # 默认只转第一个sheet
    sheet_name = book.sheetnames[0]
    sheet = book[sheet_name]
    # 行数
    max_row = sheet.max_row
    # 列数
    max_col = sheet.max_column
    xml_txt = '<?xml version="1.0" encoding="utf-8">\n'
    xml_txt += "<root>\n"
    # 属性名
    prop_names = {}
    # 一行一行遍历
    for row in range(max_row):
        row += 1
        if 2 == row:
            for col in range(max_col):
                col += 1
                prop_names[col] = sheet.cell(row, col).value
        elif 3 <= row:
            xml_line = "    <item "
            for col in range(max_col):
                col += 1
                value = sheet.cell(row, col).value
                value = "" if None == value else value
                xml_line += '%s="%s" '%(prop_names[col], value)
            xml_line += '/>\n'
            xml_txt += xml_line
    xml_txt += "</root>" 
    # 保存为xml
    xml_dir = os.path.dirname(xml_path)
    if not exists(xml_dir):
        os.makedirs(xml_dir)
    xml_f = open(xml_path, 'w', encoding='utf-8')
    xml_f.write(xml_txt)
    xml_f.close()

if '__main__' == __name__:
    excel2xml(u"Excels/测试.xlsx", "./output/test.xml")
    excel2xml(u"Excels/测试2.xlsx", "./output/test2.xml")
        