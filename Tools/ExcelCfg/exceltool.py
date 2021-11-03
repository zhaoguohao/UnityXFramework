from genericpath import exists
import openpyxl
import os
import json2lua
import json

# 遍历excel的单元格
def walk_excel_cell(excel_path):
    book = openpyxl.load_workbook(excel_path)
    # 默认只转第一个sheet
    sheet_name = book.sheetnames[0]
    sheet = book[sheet_name]
    # 行数
    max_row = sheet.max_row
    # 列数
    max_col = sheet.max_column

    # 属性名
    field_names = {}
    # 数据类型
    data_types = {}
    # 一行一行遍历
    for row in range(max_row):
        row += 1
        if 2 == row:
            for col in range(max_col):
                col += 1
                field_names[col] = sheet.cell(row, col).value
        elif 3 == row:
            for col in range(max_col):
                col += 1
                data_types[col] = sheet.cell(row, col).value
        elif 4 <= row:
            for col in range(max_col):
                col += 1
                value = sheet.cell(row, col).value
                value = "" if None == value else value
                yield col, max_col, field_names[col], data_types[col], value


def save_cfg(f_path, f_content):
    f_dir = os.path.dirname(f_path)
    if not exists(f_dir):
        os.makedirs(f_dir)
    f = open(f_path, 'w', encoding='utf-8')
    f.write(f_content)
    f.close()


def excel2xml(excel_path, xml_path):
    xml_txt = '<?xml version="1.0" encoding="utf-8">\n'
    xml_txt += "<root>\n"
    xml_line = ""
    for col, max_col, field_name, _, value in walk_excel_cell(excel_path):
        if 1 == col:
            xml_line = '    <item %s="%s"' % (field_name, value)
        elif col == max_col:
            xml_line += ' %s="%s"/>\n' % (field_name, value)
            xml_txt += xml_line
        else:
            xml_line += ' %s="%s"' % (field_name, value)
    xml_txt += "</root>"
    # 保存为xml
    save_cfg(xml_path, xml_txt)
    print('excel2xml:', excel_path, '--->', xml_path)


def excel2lua(excel_path, lua_path):
    li = []
    item = None
    for col, max_col, field_name, data_type, value in walk_excel_cell(excel_path):
        if 1 == col:
            item = {}
            item[field_name] = (data_type, value)
        elif col == max_col:
            item[field_name] = (data_type, value)
            li.append(item)
        else:
            item[field_name] = (data_type, value)
    lua_txt = 'return\n' + json2lua.dic_to_lua_str(li)
    # 保存为lua
    save_cfg(lua_path, lua_txt)
    print('excel2lua:', excel_path, '--->', lua_path)

def excel2json(excel_path, json_path):
    li = []
    item = None
    for col, max_col, field_name, _, value in walk_excel_cell(excel_path):
        if 1 == col:
            item = {}
            item[field_name] = value
        elif col == max_col:
            item[field_name] = value
            li.append(item)
        else:
            item[field_name] = value
    # 保存为lua
    json_txt = json.dumps(li, ensure_ascii=False, sort_keys=True, indent=4)
    save_cfg(json_path, json_txt)
    print('excel2json:', excel_path, '--->', json_path)